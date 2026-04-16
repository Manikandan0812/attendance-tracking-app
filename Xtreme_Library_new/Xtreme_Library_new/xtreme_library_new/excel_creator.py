import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, PatternFill
import logging

def create_excel(data_list_dict, excel_file, sheet_name):
    """
    Creates an Excel file and writes the provided test case data into the specified sheet.

    Args:
        data_list_dict (list): List of dictionaries containing test case metadata.
        excel_file (str): Path to the Excel file to create or overwrite.
        sheet_name (str): Name of the sheet in the Excel file.

    Returns:
        str: Path to the generated Excel file.
    """
    try:
        df = pd.DataFrame(data_list_dict)
        df.to_excel(excel_file, sheet_name=sheet_name, index=False)

        wb = load_workbook(excel_file)
        ws = wb.active

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
                cell.border = thin_border

        green_fill = PatternFill(start_color="6CC644", end_color="6CC644", fill_type="solid")
        red_fill = PatternFill(start_color="ED6C63", end_color="ED6C63", fill_type="solid")

        for row in range(2, len(df) + 2):
            try:
                status_cell = ws.cell(row=row, column=6)
                main_status_cell = ws.cell(row=row, column=7)

                status = str(status_cell.value).upper() if status_cell.value else ""
                main_status = str(main_status_cell.value).upper() if main_status_cell.value else ""

                if "PASS" in status or "PASS" in main_status:
                    status_cell.fill = green_fill
                    main_status_cell.fill = green_fill
                elif "FAIL" in status or "FAIL" in main_status:
                    status_cell.fill = red_fill
                    main_status_cell.fill = red_fill

                # Video hyperlink
                video_path_cell = ws.cell(row=row, column=9)
                video_path = video_path_cell.value
                step_number = ws.cell(row=row, column=2).value or row - 1
                display_text = f"video_{step_number}"

                if video_path:
                    logging.info("video_path: %s", video_path)
                    video_path_cell.value = f'=HYPERLINK("{video_path}", "{display_text}")'

                # Screenshot hyperlink
                screenshot_path_cell = ws.cell(row=row, column=10)
                screenshot_path = screenshot_path_cell.value
                screenshot_display_text = f"screenshot_{step_number}"

                if screenshot_path:
                    screenshot_path_cell.value = (
                        f'=HYPERLINK("{screenshot_path}", "{screenshot_display_text}")'
                    )
            except Exception as row_err:
                logging.warning(f"Error processing row {row}: {row_err}")

        wb.save(excel_file)
        print(f"Test results saved to {excel_file}")
        return excel_file

    except Exception as e:
        logging.error(f"Failed to create Excel file: {e}")
        print(f"Error occurred while generating Excel: {e}")
        return None



